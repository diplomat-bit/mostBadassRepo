// REPOSITORY SOURCE: diplomat-bit/Fuckyou | PATH: diplomat-bit-Fuckyou-70f83c5/apps/camt053_transaction_exporter/exporter.py
================================================================================

import csv
import json
import io
from typing import List, Dict, Any, Union, Optional
from dataclasses import dataclass, field, asdict
from datetime import datetime, date
from decimal import Decimal

# Try to import openpyxl for Excel export, fallback to CSV-based Excel if not available
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ==========================================
# 1. DATA MODELS FOR PARSED CAMT.053
# ==========================================

@dataclass
class CamtTransaction:
    """Represents a single transaction entry (Ntry/NtryDtls) in CAMT.053."""
    entry_reference: Optional[str] = None
    instruction_id: Optional[str] = None
    end_to_end_id: Optional[str] = None
    uetr: Optional[str] = None
    amount: Decimal = Decimal("0.00")
    currency: str = "EUR"
    credit_debit_indicator: str = "CRDT"  # CRDT (Credit/Inflow) or DBIT (Debit/Outflow)
    status: str = "BOOK"  # BOOK (Booked) or PDNG (Pending)
    booking_date: Optional[date] = None
    value_date: Optional[date] = None
    
    # Bank Transaction Code (ISO 20022)
    domain_code: Optional[str] = None
    family_code: Optional[str] = None
    sub_family_code: Optional[str] = None
    proprietary_code: Optional[str] = None
    
    # Counterparty Details
    counterparty_name: Optional[str] = None
    counterparty_iban: Optional[str] = None
    counterparty_bic: Optional[str] = None
    
    # Remittance Information
    remittance_info: Optional[str] = None
    additional_entry_info: Optional[str] = None


@dataclass
class CamtStatement:
    """Represents a statement (Stmt) in CAMT.053."""
    statement_id: str
    creation_date_time: Optional[datetime] = None
    account_iban: Optional[str] = None
    account_currency: str = "EUR"
    opening_balance_amount: Decimal = Decimal("0.00")
    opening_balance_date: Optional[date] = None
    closing_balance_amount: Decimal = Decimal("0.00")
    closing_balance_date: Optional[date] = None
    transactions: List[CamtTransaction] = field(default_factory=list)


@dataclass
class CamtDocument:
    """Represents the entire parsed CAMT.053 document."""
    message_id: str
    creation_date_time: datetime
    statements: List[CamtStatement] = field(default_factory=list)


# ==========================================
# BASE EXPORTER CLASS
# ==========================================

class BaseExporter:
    """Base class providing common export utilities and interface definitions."""
    
    def __init__(self, document: CamtDocument):
        self.document = document

    def _serialize_value(self, val: Any) -> Any:
        """Helper to serialize complex types for JSON/CSV output."""
        if isinstance(val, (datetime, date)):
            return val.isoformat()
        if isinstance(val, Decimal):
            return float(val)
        return val

    def to_csv(self) -> str:
        """Generates a CSV string of the exported data."""
        raise NotImplementedError("Subclasses must implement to_csv")

    def to_json(self, indent: int = 2) -> str:
        """Generates a JSON string of the exported data."""
        raise NotImplementedError("Subclasses must implement to_json")

    def to_xlsx(self) -> bytes:
        """Generates an Excel workbook as bytes."""
        raise NotImplementedError("Subclasses must implement to_xlsx")


# ==========================================
# APP 1: STANDARD TRANSACTION LEDGER EXPORTER
# ==========================================

class StandardLedgerExporter(BaseExporter):
    """
    App 1: Standard Transaction Ledger Exporter
    Focuses on standard accounting and reconciliation columns.
    Maps ISO 20022 elements to standard ledger fields.
    """
    
    HEADERS = [
        "Statement ID", "Booking Date", "Value Date", "Direction", 
        "Amount", "Currency", "Counterparty Name", "Counterparty IBAN", 
        "Remittance Info", "Reference ID", "Status"
    ]

    def _get_rows(self) -> List[Dict[str, Any]]:
        rows = []
        for stmt in self.document.statements:
            for tx in stmt.transactions:
                direction = "Inflow" if tx.credit_debit_indicator == "CRDT" else "Outflow"
                # For standard ledger, represent outflows as negative numbers
                signed_amount = tx.amount if tx.credit_debit_indicator == "CRDT" else -tx.amount
                
                rows.append({
                    "Statement ID": stmt.statement_id,
                    "Booking Date": tx.booking_date,
                    "Value Date": tx.value_date,
                    "Direction": direction,
                    "Amount": signed_amount,
                    "Currency": tx.currency,
                    "Counterparty Name": tx.counterparty_name or "Unknown",
                    "Counterparty IBAN": tx.counterparty_iban or "",
                    "Remittance Info": tx.remittance_info or "",
                    "Reference ID": tx.end_to_end_id or tx.entry_reference or "",
                    "Status": tx.status
                })
        return rows

    def to_csv(self) -> str:
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=self.HEADERS)
        writer.writeheader()
        for row in self._get_rows():
            serialized_row = {k: self._serialize_value(v) for k, v in row.items()}
            writer.writerow(serialized_row)
        return output.getvalue()

    def to_json(self, indent: int = 2) -> str:
        rows = self._get_rows()
        serialized_rows = []
        for row in rows:
            serialized_rows.append({k: self._serialize_value(v) for k, v in row.items()})
        return json.dumps(serialized_rows, indent=indent)

    def to_xlsx(self) -> bytes:
        if not HAS_OPENPYXL:
            # Fallback to CSV if openpyxl is not installed
            return self.to_csv().encode('utf-8')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transaction Ledger"
        
        # Styling
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center")
        right_align = Alignment(horizontal="right")
        
        # Write Headers
        ws.append(self.HEADERS)
        for col_num, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Write Rows
        for row_idx, row in enumerate(self._get_rows(), 2):
            ws.append([
                row["Statement ID"],
                row["Booking Date"],
                row["Value Date"],
                row["Direction"],
                float(row["Amount"]),
                row["Currency"],
                row["Counterparty Name"],
                row["Counterparty IBAN"],
                row["Remittance Info"],
                row["Reference ID"],
                row["Status"]
            ])
            
            # Format specific columns
            ws.cell(row=row_idx, column=2).number_format = 'yyyy-mm-dd'
            ws.cell(row=row_idx, column=3).number_format = 'yyyy-mm-dd'
            ws.cell(row=row_idx, column=4).alignment = center_align
            
            # Format Amount
            amt_cell = ws.cell(row=row_idx, column=5)
            amt_cell.number_format = '#,##0.00'
            amt_cell.alignment = right_align
            if row["Amount"] < 0:
                amt_cell.font = Font(color="9C0006") # Red text for negative amounts

        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()


# ==========================================
# APP 2: TREASURY CASH FLOW SUMMARY EXPORTER
# ==========================================

class TreasurySummaryExporter(BaseExporter):
    """
    App 2: Treasury Cash Flow Summary Exporter
    Aggregates transactions by date, currency, and direction to provide a high-level cash flow summary.
    """
    
    HEADERS = ["Date", "Currency", "Total Inflow", "Total Outflow", "Net Cash Flow", "Transaction Count"]

    def _get_summary_data(self) -> List[Dict[str, Any]]:
        summary = {}
        for stmt in self.document.statements:
            for tx in stmt.transactions:
                tx_date = tx.booking_date or tx.value_date or date.today()
                key = (tx_date, tx.currency)
                
                if key not in summary:
                    summary[key] = {
                        "Inflow": Decimal("0.00"),
                        "Outflow": Decimal("0.00"),
                        "Count": 0
                    }
                
                if tx.credit_debit_indicator == "CRDT":
                    summary[key]["Inflow"] += tx.amount
                else:
                    summary[key]["Outflow"] += tx.amount
                summary[key]["Count"] += 1

        rows = []
        for (tx_date, currency), metrics in sorted(summary.items(), key=lambda x: x[0]):
            net_flow = metrics["Inflow"] - metrics["Outflow"]
            rows.append({
                "Date": tx_date,
                "Currency": currency,
                "Total Inflow": metrics["Inflow"],
                "Total Outflow": metrics["Outflow"],
                "Net Cash Flow": net_flow,
                "Transaction Count": metrics["Count"]
            })
        return rows

    def to_csv(self) -> str:
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=self.HEADERS)
        writer.writeheader()
        for row in self._get_summary_data():
            serialized_row = {k: self._serialize_value(v) for k, v in row.items()}
            writer.writerow(serialized_row)
        return output.getvalue()

    def to_json(self, indent: int = 2) -> str:
        rows = self._get_summary_data()
        serialized_rows = []
        for row in rows:
            serialized_rows.append({k: self._serialize_value(v) for k, v in row.items()})
        return json.dumps(serialized_rows, indent=indent)

    def to_xlsx(self) -> bytes:
        if not HAS_OPENPYXL:
            return self.to_csv().encode('utf-8')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Treasury Cash Flow Summary"
        
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        center_align = Alignment(horizontal="center")
        right_align = Alignment(horizontal="right")
        
        ws.append(self.HEADERS)
        for col_num, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for row_idx, row in enumerate(self._get_summary_data(), 2):
            ws.append([
                row["Date"],
                row["Currency"],
                float(row["Total Inflow"]),
                float(row["Total Outflow"]),
                float(row["Net Cash Flow"]),
                row["Transaction Count"]
            ])
            
            ws.cell(row=row_idx, column=1).number_format = 'yyyy-mm-dd'
            ws.cell(row=row_idx, column=2).alignment = center_align
            
            # Format numeric columns
            for col_num in [3, 4, 5]:
                cell = ws.cell(row=row_idx, column=col_num)
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
                if col_num == 5 and row["Net Cash Flow"] < 0:
                    cell.font = Font(color="9C0006")
            
            ws.cell(row=row_idx, column=6).alignment = right_align

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()


# ==========================================
# APP 3: AUDIT & COMPLIANCE DETAILED EXPORTER
# ==========================================

class AuditComplianceExporter(BaseExporter):
    """
    App 3: Audit & Compliance Detailed Exporter
    Exports every single ISO 20022 field, including proprietary codes, entry references,
    status, and unique identifiers (UETR, EndToEndID) for deep auditing.
    """
    
    HEADERS = [
        "Message ID", "Statement ID", "Account IBAN", "Booking Date", "Value Date",
        "Amount", "Currency", "Indicator", "Status", "Entry Ref", "Instruction ID",
        "End To End ID", "UETR", "Domain Code", "Family Code", "SubFamily Code",
        "Proprietary Code", "Counterparty Name", "Counterparty IBAN", "Counterparty BIC",
        "Remittance Info", "Additional Info"
    ]

    def _get_audit_rows(self) -> List[Dict[str, Any]]:
        rows = []
        for stmt in self.document.statements:
            for tx in stmt.transactions:
                rows.append({
                    "Message ID": self.document.message_id,
                    "Statement ID": stmt.statement_id,
                    "Account IBAN": stmt.account_iban or "",
                    "Booking Date": tx.booking_date,
                    "Value Date": tx.value_date,
                    "Amount": tx.amount,
                    "Currency": tx.currency,
                    "Indicator": tx.credit_debit_indicator,
                    "Status": tx.status,
                    "Entry Ref": tx.entry_reference or "",
                    "Instruction ID": tx.instruction_id or "",
                    "End To End ID": tx.end_to_end_id or "",
                    "UETR": tx.uetr or "",
                    "Domain Code": tx.domain_code or "",
                    "Family Code": tx.family_code or "",
                    "SubFamily Code": tx.sub_family_code or "",
                    "Proprietary Code": tx.proprietary_code or "",
                    "Counterparty Name": tx.counterparty_name or "",
                    "Counterparty IBAN": tx.counterparty_iban or "",
                    "Counterparty BIC": tx.counterparty_bic or "",
                    "Remittance Info": tx.remittance_info or "",
                    "Additional Info": tx.additional_entry_info or ""
                })
        return rows

    def to_csv(self) -> str:
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=self.HEADERS)
        writer.writeheader()
        for row in self._get_audit_rows():
            serialized_row = {k: self._serialize_value(v) for k, v in row.items()}
            writer.writerow(serialized_row)
        return output.getvalue()

    def to_json(self, indent: int = 2) -> str:
        rows = self._get_audit_rows()
        serialized_rows = []
        for row in rows:
            serialized_rows.append({k: self._serialize_value(v) for k, v in row.items()})
        return json.dumps(serialized_rows, indent=indent)

    def to_xlsx(self) -> bytes:
        if not HAS_OPENPYXL:
            return self.to_csv().encode('utf-8')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit & Compliance Log"
        
        header_font = Font(name="Consolas", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")
        center_align = Alignment(horizontal="center")
        right_align = Alignment(horizontal="right")
        
        ws.append(self.HEADERS)
        for col_num, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for row_idx, row in enumerate(self._get_audit_rows(), 2):
            ws.append([
                row["Message ID"],
                row["Statement ID"],
                row["Account IBAN"],
                row["Booking Date"],
                row["Value Date"],
                float(row["Amount"]),
                row["Currency"],
                row["Indicator"],
                row["Status"],
                row["Entry Ref"],
                row["Instruction ID"],
                row["End To End ID"],
                row["UETR"],
                row["Domain Code"],
                row["Family Code"],
                row["SubFamily Code"],
                row["Proprietary Code"],
                row["Counterparty Name"],
                row["Counterparty IBAN"],
                row["Counterparty BIC"],
                row["Remittance Info"],
                row["Additional Info"]
            ])
            
            ws.cell(row=row_idx, column=4).number_format = 'yyyy-mm-dd'
            ws.cell(row=row_idx, column=5).number_format = 'yyyy-mm-dd'
            ws.cell(row=row_idx, column=6).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=6).alignment = right_align
            ws.cell(row=row_idx, column=8).alignment = center_align
            ws.cell(row=row_idx, column=9).alignment = center_align

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()


# ==========================================
# APP 4: CUSTOMER/VENDOR REMITTANCE MATCHER
# ==========================================

class RemittanceMatcherExporter(BaseExporter):
    """
    App 4: Customer/Vendor Remittance Matcher Exporter
    Focuses on matching remittance info, debtor/creditor names, and accounts.
    Optimized for ERP ingestion and automated reconciliation matching.
    """
    
    HEADERS = [
        "Match Key", "Counterparty Name", "Remittance Info", "Amount", 
        "Currency", "Direction", "End To End ID", "UETR", "Account IBAN"
    ]

    def _get_matcher_rows(self) -> List[Dict[str, Any]]:
        rows = []
        for stmt in self.document.statements:
            for tx in stmt.transactions:
                # Generate a unique match key for ERP systems
                match_key = tx.end_to_end_id or tx.uetr or tx.entry_reference or "N/A"
                direction = "RECEIPT" if tx.credit_debit_indicator == "CRDT" else "PAYMENT"
                
                rows.append({
                    "Match Key": match_key,
                    "Counterparty Name": tx.counterparty_name or "Unknown Counterparty",
                    "Remittance Info": tx.remittance_info or "",
                    "Amount": tx.amount,
                    "Currency": tx.currency,
                    "Direction": direction,
                    "End To End ID": tx.end_to_end_id or "",
                    "UETR": tx.uetr or "",
                    "Account IBAN": tx.counterparty_iban or ""
                })
        return rows

    def to_csv(self) -> str:
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=self.HEADERS)
        writer.writeheader()
        for row in self._get_matcher_rows():
            serialized_row = {k: self._serialize_value(v) for k, v in row.items()}
            writer.writerow(serialized_row)
        return output.getvalue()

    def to_json(self, indent: int = 2) -> str:
        rows = self._get_matcher_rows()
        serialized_rows = []
        for row in rows:
            serialized_rows.append({k: self._serialize_value(v) for k, v in row.items()})
        return json.dumps(serialized_rows, indent=indent)

    def to_xlsx(self) -> bytes:
        if not HAS_OPENPYXL:
            return self.to_csv().encode('utf-8')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ERP Remittance Matcher"
        
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        center_align = Alignment(horizontal="center")
        right_align = Alignment(horizontal="right")
        
        ws.append(self.HEADERS)
        for col_num, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for row_idx, row in enumerate(self._get_matcher_rows(), 2):
            ws.append([
                row["Match Key"],
                row["Counterparty Name"],
                row["Remittance Info"],
                float(row["Amount"]),
                row["Currency"],
                row["Direction"],
                row["End To End ID"],
                row["UETR"],
                row["Account IBAN"]
            ])
            
            ws.cell(row=row_idx, column=4).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=4).alignment = right_align
            ws.cell(row=row_idx, column=5).alignment = center_align
            ws.cell(row=row_idx, column=6).alignment = center_align

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()


# ==========================================
# UNIFIED EXPORT ENGINE ORCHESTRATOR
# ==========================================

class CamtExporterEngine:
    """
    Unified engine to run any of the 4 export apps with specified formats.
    """
    
    APPS = {
        "ledger": StandardLedgerExporter,
        "treasury": TreasurySummaryExporter,
        "audit": AuditComplianceExporter,
        "matcher": RemittanceMatcherExporter
    }

    @classmethod
    def export(cls, document: CamtDocument, app_name: str, format_type: str) -> Union[str, bytes]:
        """
        Exports the CAMT document using the specified app and format.
        
        :param document: Parsed CamtDocument object.
        :param app_name: One of 'ledger', 'treasury', 'audit', 'matcher'.
        :param format_type: One of 'csv', 'json', 'xlsx'.
        :return: String (for CSV/JSON) or Bytes (for XLSX).
        """
        app_class = cls.APPS.get(app_name.lower())
        if not app_class:
            raise ValueError(f"Unknown app name: {app_name}. Choose from {list(cls.APPS.keys())}")
        
        exporter = app_class(document)
        
        fmt = format_type.lower()
        if fmt == "csv":
            return exporter.to_csv()
        elif fmt == "json":
            return exporter.to_json()
        elif fmt == "xlsx":
            return exporter.to_xlsx()
        else:
            raise ValueError(f"Unsupported format: {format_type}. Choose from 'csv', 'json', 'xlsx'")


# ==========================================
# DEMO / TEST GENERATOR
# ==========================================

def generate_mock_camt_data() -> CamtDocument:
    """Generates a mock CamtDocument for testing and demonstration purposes."""
    tx1 = CamtTransaction(
        entry_reference="TX-2023-001",
        instruction_id="INS-99281",
        end_to_end_id="E2E-8829102",
        uetr: "f81d4fae-7dec-11d0-a765-00a0c91e6bf6",
        amount=Decimal("1500.50"),
        currency="EUR",
        credit_debit_indicator="CRDT",
        status="BOOK",
        booking_date=date(2023, 10, 1),
        value_date=date(2023, 10, 1),
        domain_code="PMNT",
        family_code="RCDT",
        sub_family_code="MCT",
        proprietary_code="PROP-112",
        counterparty_name="ACME Corporation",
        counterparty_iban="DE89370400440532013000",
        counterparty_bic="DBEUMM21XXX",
        remittance_info="Invoice INV-2023-9910",
        additional_entry_info="Urgent processing requested"
    )

    tx2 = CamtTransaction(
        entry_reference="TX-2023-002",
        instruction_id="INS-99282",
        end_to_end_id="E2E-8829103",
        uetr: "a42d4fae-7dec-11d0-a765-00a0c91e6bf6",
        amount=Decimal("450.00"),
        currency="EUR",
        credit_debit_indicator="DBIT",
        status="BOOK",
        booking_date=date(2023, 10, 2),
        value_date=date(2023, 10, 2),
        domain_code="PMNT",
        family_code="RCDT",
        sub_family_code="MCT",
        proprietary_code="PROP-113",
        counterparty_name="Office Supplies Depot",
        counterparty_iban="DE89370400440532013999",
        counterparty_bic="DBEUMM21XXX",
        remittance_info="Stationery and paper supplies",
        additional_entry_info=""
    )

    stmt = CamtStatement(
        statement_id="STMT-2023-10",
        creation_date_time=datetime(2023, 10, 3, 12, 0, 0),
        account_iban="DE89370400440532011111",
        account_currency="EUR",
        opening_balance_amount=Decimal("10000.00"),
        opening_balance_date=date(2023, 10, 1),
        closing_balance_amount=Decimal("11050.50"),
        closing_balance_date=date(2023, 10, 3),
        transactions=[tx1, tx2]
    )

    return CamtDocument(
        message_id="MSG-99281029",
        creation_date_time=datetime(2023, 10, 3, 12, 5, 0),
        statements=[stmt]
    )


if __name__ == "__main__":
    # Quick self-test and demonstration of the 4 apps
    doc = generate_mock_camt_data()
    
    print("=== APP 1: Standard Ledger (CSV) ===")
    print(CamtExporterEngine.export(doc, "ledger", "csv"))
    
    print("\n=== APP 2: Treasury Summary (JSON) ===")
    print(CamtExporterEngine.export(doc, "treasury", "json"))
    
    print("\n=== APP 3: Audit & Compliance (CSV) ===")
    print(CamtExporterEngine.export(doc, "audit", "csv")[:500] + "...\n[TRUNCATED]")
    
    print("\n=== APP 4: Remittance Matcher (JSON) ===")
    print(CamtExporterEngine.export(doc, "matcher", "json"))
    
    # Test Excel generation
    try:
        xlsx_data = CamtExporterEngine.export(doc, "ledger", "xlsx")
        print(f"\nExcel generation successful! Size: {len(xlsx_data)} bytes")
    except Exception as e:
        print(f"\nExcel generation failed or skipped: {e}")